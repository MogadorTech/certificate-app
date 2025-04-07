

from flask import Flask, render_template, request, send_file
import os
import hashlib
import qrcode
import fitz  # PyMuPDF
import io
import openpyxl
from datetime import datetime
import uuid



app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Route for the home page
@app.route('/')
def home():
    return render_template('home.html', now=datetime.now())

# Route for University Administrator (Upload Certificate)
@app.route('/admin')
def admin():
    return render_template('admin.html', now=datetime.now())


# Route for uploading and modifying the certificate
@app.route('/upload', methods=['POST'])
def upload():
    uploaded_file = request.files['pdf']
    if uploaded_file.filename.endswith('.pdf'):
        input_path = os.path.join(UPLOAD_FOLDER, uploaded_file.filename)
        output_path = os.path.join(OUTPUT_FOLDER, 'modified_' + uploaded_file.filename)
        uploaded_file.save(input_path)

        # Hash the PDF
        with open(input_path, 'rb') as f:
            file_bytes = f.read()
            pdf_hash = hashlib.sha256(file_bytes).hexdigest()

        # Create QR code
        verification_url = f"https://yourappname.onrender.com/verify?hash={pdf_hash}"
        qr = qrcode.make(verification_url)
        qr_io = io.BytesIO()
        qr.save(qr_io, format='PNG')
        qr_io.seek(0)

        # Modify PDF
        doc = fitz.open(input_path)
        page = doc[0]
        qr_rect = fitz.Rect(50, 50, 200, 200)
        page.insert_image(qr_rect, stream=qr_io.read())
        text_rect = fitz.Rect(50, 210, 500, 250)
        page.insert_textbox(text_rect, f"SHA-256 Hash:\n{pdf_hash}", fontsize=8)
        doc.save(output_path)
        doc.close()

        qr_filename = f"qr_{uuid.uuid4().hex[:8]}.png"
        qr_path = os.path.join(OUTPUT_FOLDER, qr_filename)

        # Ensure OUTPUT_FOLDER exists before saving the QR image
        os.makedirs(OUTPUT_FOLDER, exist_ok=True)

        # Create QR code and save to disk
        qr = qrcode.make(pdf_hash)
        qr.save(qr_path)

        # Log the data to Excel with QR path
        log_to_excel(uploaded_file.filename, pdf_hash, qr_path)


        return send_file(output_path, as_attachment=True)

    return "Invalid file type. Please upload a PDF."

from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.exceptions import InvalidFileException

def log_to_excel(cert_name, cert_hash, qr_path):
    log_file = 'outputs/log.xlsx'

    try:
        # Create file if it doesn't exist
        if not os.path.exists(log_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Certificate Logs"
            ws.append(["ID", "Certificate Name", "SHA-256 Hash", "QR Code", "Date"])
            wb.save(log_file)

        wb = openpyxl.load_workbook(log_file)
        ws = wb.active

        cert_id = str(uuid.uuid4())[:8]
        date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Add data except QR image
        row = [cert_id, cert_name, cert_hash, "", date_str]
        ws.append(row)
        row_number = ws.max_row

        # Add QR image to the cell in the 4th column (D)
        img = XLImage(qr_path)
        img.width, img.height = 60, 60  # Resize if needed
        cell_position = f"D{row_number}"
        ws.add_image(img, cell_position)
        wb.save(log_file)

    except InvalidFileException as e:
        print(f"Error: Unable to access the Excel file. It may be open in another application. {e}")
    except PermissionError as e:
        print(f"Error: Permission denied. Please ensure the Excel file is not open in another program. {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

#route to Display a form to enter the certificate hash.
@app.route('/verify', methods=['GET', 'POST'])
def verify():
    if request.method == 'POST':
        input_hash = request.form['hash'].strip()

        log_file = 'outputs/log.xlsx'
        found = False

        if os.path.exists(log_file):
            wb = openpyxl.load_workbook(log_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[2] == input_hash:  # Column C is the hash
                    found = True
                    break

        return render_template('verify_result.html', hash_value=input_hash, found=found, now=datetime.now())

    return render_template('verify.html', now=datetime.now())


#function to verify certificate based on certificate hash
def verify_certificate(certificate_hash):
    log_file = 'outputs/log.xlsx'

    if not os.path.exists(log_file):
        return False, None, None

    wb = openpyxl.load_workbook(log_file)
    ws = wb.active

    # Loop through rows to find the certificate with the matching hash
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        stored_hash = row[2]  # The hash is stored in the 3rd column (C)
        if stored_hash == certificate_hash:
            cert_name = row[1]  # Certificate name is in the 2nd column (B)
            cert_date = row[4]  # Date is in the 5th column (E)
            return True, cert_name, cert_date

    return False, None, None

if __name__ == '__main__':
    app.run(debug=True)
